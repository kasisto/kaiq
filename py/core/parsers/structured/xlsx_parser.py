# type: ignore
import logging
from io import BytesIO
from typing import AsyncGenerator

import networkx as nx
import numpy as np
from openpyxl import load_workbook

from core.base.parsers.base_parser import AsyncParser
from core.base.providers import (
    CompletionProvider,
    DatabaseProvider,
    IngestionConfig,
)

logger = logging.getLogger(__name__)


class XLSXParser(AsyncParser[str | bytes]):
    """A parser for XLSX data."""

    def __init__(
        self,
        config: IngestionConfig,
        database_provider: DatabaseProvider,
        llm_provider: CompletionProvider,
    ):
        self.database_provider = database_provider
        self.llm_provider = llm_provider
        self.config = config
        self.load_workbook = load_workbook

    async def ingest(
        self, data: bytes, *args, **kwargs
    ) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each row."""
        if isinstance(data, str):
            raise ValueError("XLSX data must be in bytes format.")

        wb = self.load_workbook(filename=BytesIO(data))
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                yield ", ".join(map(str, row))


class XLSXParserAdvanced(AsyncParser[str | bytes]):
    """A parser for XLSX data."""

    # identifies connected components in the excel graph and extracts data from each component
    def __init__(
        self,
        config: IngestionConfig,
        database_provider: DatabaseProvider,
        llm_provider: CompletionProvider,
        ocr_provider=None,
    ):
        self.database_provider = database_provider
        self.llm_provider = llm_provider
        self.config = config
        self.nx = nx
        self.np = np
        self.load_workbook = load_workbook

    def connected_components(self, arr):
        if arr.size == 0:
            return
        g = self.nx.grid_2d_graph(len(arr), len(arr[0]))
        # Use vectorized None check - create boolean mask for empty cells
        # arr == None works element-wise for object arrays
        empty_mask = self.np.vectorize(lambda x: x is None)(arr)
        empty_cell_indices = list(
            zip(*self.np.where(empty_mask), strict=False)
        )
        g.remove_nodes_from(empty_cell_indices)
        components = self.nx.connected_components(g)
        for component in components:
            if not component:
                continue
            rows, cols = zip(*component, strict=False)
            min_row, max_row = min(rows), max(rows)
            min_col, max_col = min(cols), max(cols)
            yield arr[min_row : max_row + 1, min_col : max_col + 1].astype(
                "str"
            )

    async def ingest(
        self, data: bytes, num_col_times_num_rows: int = 100, *args, **kwargs
    ) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each connected component."""
        if isinstance(data, str):
            raise ValueError("XLSX data must be in bytes format.")

        workbook = self.load_workbook(filename=BytesIO(data))
        chunks_yielded = 0

        for ws in workbook.worksheets:
            logger.debug(f"Processing worksheet: {ws.title}")
            ws_data = self.np.array(
                [[cell.value for cell in row] for row in ws.iter_rows()]
            )
            logger.debug(f"Worksheet {ws.title} has shape: {ws_data.shape}")

            if ws_data.size == 0:
                logger.warning(f"Worksheet {ws.title} is empty, skipping")
                continue

            table_count = 0
            for table in self.connected_components(ws_data):
                table_count += 1
                logger.debug(f"Found table {table_count} with shape: {table.shape}")

                # If table only has 1 row (just headers or single data), yield it directly
                if len(table) == 1:
                    row_text = ", ".join(table[0])
                    if row_text.strip():
                        logger.debug(f"Yielding single-row table: {row_text[:100]}...")
                        yield row_text
                        chunks_yielded += 1
                    continue

                num_rows = len(table)
                num_rows_per_chunk = max(1, num_col_times_num_rows // num_rows)
                headers = ", ".join(table[0])

                # add header to each chunk
                for i in range(1, num_rows, num_rows_per_chunk):
                    chunk = table[i : i + num_rows_per_chunk]
                    if len(chunk) == 0:
                        continue
                    chunk_text = (
                        headers
                        + "\n"
                        + "\n".join([", ".join(row) for row in chunk])
                    )
                    logger.debug(f"Yielding chunk {chunks_yielded + 1}: {chunk_text[:100]}...")
                    yield chunk_text
                    chunks_yielded += 1

            if table_count == 0:
                logger.warning(f"No connected components found in worksheet {ws.title}")

        logger.info(f"XLSXParserAdvanced completed: yielded {chunks_yielded} chunks")
